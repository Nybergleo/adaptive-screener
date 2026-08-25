"""Local Adaptive Screener Excel workbook generator.

Run this script locally, open the served page, choose fields and filters, then
download an Excel workbook that stores the setup and generated BQL query.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
import argparse
import csv
import html
import json
import os
from pathlib import Path
import re
import urllib.parse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import xlsxwriter


WESTERN_EUROPE_COUNTRIES = {
    "Austria": "AT",
    "Belgium": "BE",
    "Denmark": "DK",
    "Finland": "FI",
    "France": "FR",
    "Germany": "DE",
    "Ireland": "IE",
    "Italy": "IT",
    "Luxembourg": "LU",
    "Netherlands": "NL",
    "Norway": "NO",
    "Portugal": "PT",
    "Spain": "ES",
    "Sweden": "SE",
    "Switzerland": "CH",
    "United Kingdom": "GB",
}

NORTH_AMERICA_COUNTRIES = {
    "Canada": "CA",
    "United States": "US",
}

COUNTRY_GROUPS = {
    "western_europe": WESTERN_EUROPE_COUNTRIES,
    "north_america": NORTH_AMERICA_COUNTRIES,
}


@dataclass(frozen=True)
class FieldSpec:
    key: str
    label: str
    alias: str
    expression: str
    field_type: str
    definition: str
    default_min: str = ""
    default_max: str = ""
    default_text: str = ""
    default_input: bool = True
    default_output: bool = True
    format_type: str = "number"

    def expression_for(self, currency: str, n_quarters: int) -> str:
        fpo_start = -(int(n_quarters) - 1)
        return self.expression.format(currency=currency, fpo_start=fpo_start, n_quarters=n_quarters)


IDENTITY_FIELDS = [
    FieldSpec(
        "name",
        "Company name",
        "name",
        "name()",
        "text",
        "Input: security identifier. Output units: string company/security name.",
        default_input=False,
    ),
    FieldSpec(
        "cntry_of_domicile",
        "Country of domicile",
        "cntry_of_domicile",
        "cntry_of_domicile()",
        "text",
        "Input: security identifier. Output units: string ISO-style country code for country of domicile.",
        default_input=False,
    ),
    FieldSpec(
        "crncy",
        "Trading currency",
        "crncy",
        "crncy()",
        "text",
        "Input: security identifier. Output units: string ISO currency code.",
        default_input=False,
    ),
    FieldSpec(
        "sector",
        "BICS sector",
        "sector",
        "classification_name(BICS, 1)",
        "text",
        "Input: security identifier plus BICS taxonomy level 1. Output units: string sector name.",
        default_input=False,
    ),
    FieldSpec(
        "industry_group",
        "BICS industry group",
        "industry_group",
        "classification_name(BICS, 2)",
        "text",
        "Input: security identifier plus BICS taxonomy level 2. Output units: string industry-group name.",
        default_input=False,
    ),
]


FIELD_CATALOG = [
    FieldSpec(
        "ebit_margin_ltm",
        "LTM EBIT / operating margin",
        "ebit_margin_ltm",
        "oper_margin(fpt=LTM)",
        "numeric",
        "Input: security identifier plus fpt=LTM period selector. Output units: decimal ratio (0.25 = 25%). Min/Max filters apply to this decimal output.",
        default_min="0.10",
        format_type="percent_points",
    ),
    FieldSpec(
        "ebita_margin_ltm",
        "LTM EBITA margin",
        "ebita_margin_ltm",
        "ebita_margin(fpt=LTM)",
        "numeric",
        "Input: security identifier plus fpt=LTM period selector. Output units: decimal ratio. EBITA margin is EBIT plus amortization as a share of revenue. Min/Max filters apply to this decimal output.",
        default_min="0.11",
        format_type="percent_points",
    ),
    FieldSpec(
        "ebitda_margin_ltm",
        "LTM EBITDA margin",
        "ebitda_margin_ltm",
        "ebitda_to_revenue(fpt=LTM)",
        "numeric",
        "Input: security identifier plus fpt=LTM period selector. Output units: decimal ratio. This ratio field returns EBITDA divided by revenue. Min/Max filters apply to this decimal output.",
        format_type="percent_points",
    ),
    FieldSpec(
        "margin_stability_pp",
        "Quarterly EBIT margin stability",
        "margin_stability_pp",
        "std(oper_margin(fpt=Q, fpo=range({fpo_start},0)))",
        "numeric",
        "Input: time series of quarterly oper_margin decimal ratios from fpo=range({fpo_start},0). Output units: decimal ratio standard deviation, i.e. margin percentage points as a decimal (0.04 = 4 pp). Min/Max filters apply to this scalar output. Window controls the number of quarters.",
        default_max="0.04",
        format_type="percent_points",
    ),
    FieldSpec(
        "fwd_pe_blended_12m",
        "Blended 12M forward P/E",
        "fwd_pe_blended_12m",
        "pe_ratio(fpt=BT, fpo=1)",
        "numeric",
        "Input: security identifier plus fpt=BT and fpo=1 forward period selector. Output units: dimensionless valuation multiple (22.4 = 22.4x). Min/Max filters apply to this multiple.",
        default_max="25",
    ),
    FieldSpec(
        "ebit_ltm",
        "LTM EBIT",
        "ebit_ltm",
        "ebit(fpt=LTM, currency={currency})",
        "numeric",
        "Input: security identifier plus fpt=LTM period selector and selected currency. Output units: millions of {currency}. Min/Max filters apply to this monetary output in millions.",
    ),
    FieldSpec(
        "market_cap",
        "Market capitalization",
        "market_cap",
        "cur_mkt_cap(currency={currency})",
        "numeric",
        "Input: security identifier plus selected currency. Output units: millions of {currency}. Min/Max filters apply to this monetary output in millions.",
        default_min="500",
    ),
    FieldSpec(
        "turnover",
        "Turnover",
        "turnover",
        "turnover(currency={currency})",
        "numeric",
        "Input: security identifier plus selected currency. Output units: millions of {currency}, representing daily trading value. Min/Max filters apply to this monetary output in millions.",
    ),
    FieldSpec(
        "revenue_ltm",
        "LTM revenue",
        "revenue_ltm",
        "sales_rev_turn(fpt=LTM, currency={currency})",
        "numeric",
        "Input: security identifier plus fpt=LTM period selector and selected currency. Output units: millions of {currency}. Min/Max filters apply to this revenue output in millions.",
    ),
    FieldSpec(
        "free_cash_flow_yield",
        "Free cash flow yield",
        "free_cash_flow_yield",
        "free_cash_flow_yield",
        "numeric",
        "Input: security identifier. Output units: decimal ratio (0.04 = 4%). Min/Max filters apply to this decimal yield output.",
        format_type="percent_points",
    ),
    FieldSpec(
        "dividend_yield",
        "Dividend yield",
        "dividend_yield",
        "dividend_yield",
        "numeric",
        "Input: security identifier. Output units: decimal ratio (0.015 = 1.5%). Min/Max filters apply to this decimal yield output.",
        format_type="percent_points",
    ),
    FieldSpec(
        "return_on_capital",
        "Return on capital employed",
        "return_on_capital_employed",
        "return_on_capital_employed",
        "numeric",
        "Input: security identifier. Output units: decimal ratio (3.22 = 322%). Uses return_on_capital_employed because return_on_capital does not resolve in BQL. Min/Max filters apply to this decimal output.",
        format_type="percent_points",
    ),
    FieldSpec(
        "sector_filter",
        "BICS sector",
        "sector",
        "classification_name(BICS, 1)",
        "text",
        "Input: security identifier plus BICS taxonomy level 1. Output units: string sector name. Text filters use exact equals or comma-separated in-list values.",
        default_output=False,
        format_type="text",
    ),
    FieldSpec(
        "industry_group_filter",
        "BICS industry group",
        "industry_group",
        "classification_name(BICS, 2)",
        "text",
        "Input: security identifier plus BICS taxonomy level 2. Output units: string industry-group name. Text filters use exact equals or comma-separated in-list values.",
        default_output=False,
        format_type="text",
    ),
]

DEFAULT_FIELD_KEYS = [
    "ebit_margin_ltm",
    "ebita_margin_ltm",
    "ebitda_margin_ltm",
    "margin_stability_pp",
    "fwd_pe_blended_12m",
    "ebit_ltm",
    "market_cap",
    "turnover",
]

CATALOG_BY_KEY = {field.key: field for field in FIELD_CATALOG}
DEFAULT_STABILITY_QUARTERS = 12
USER_TEMPLATE_PATH = Path(__file__).resolve().parent / "adaptive_screener_templates.json"
SETUP_SHEET = "Setup"
RAW_DATA_SHEET = "Raw Data"
SETUP_QUERY_CELL = "'Setup'!$B$10"
SETUP_SORT_FIELD_CELL = "'Setup'!$B$7"
SETUP_SORT_DIRECTION_CELL = "'Setup'!$B$8"

FIELD_TEMPLATES: list[dict] = []


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_text_values(value: object) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in re.split(r"[,;\n]", str(value)) if part.strip()]


def field_supports_window(expression: str) -> bool:
    return "{fpo_start}" in expression or "{n_quarters}" in expression


def field_uses_placeholders(expression: str) -> bool:
    return any(token in expression for token in ("{currency}", "{fpo_start}", "{n_quarters}"))


def bql_quote(value: str) -> str:
    return "'" + value.replace("'", "\\'") + "'"


def bql_number(value: float) -> str:
    text = f"{value:.12f}".rstrip("0").rstrip(".")
    if text == "-0":
        return "0"
    return text


def add_numeric_conditions(conditions: list[str], expression: str, minimum: object, maximum: object) -> None:
    min_value = parse_number(minimum)
    max_value = parse_number(maximum)
    if min_value is not None:
        conditions.append(f"{expression} >= {bql_number(min_value)}")
    if max_value is not None:
        conditions.append(f"{expression} <= {bql_number(max_value)}")


def normalize_field(raw: dict, currency: str) -> dict:
    key = str(raw.get("key", "")).strip()
    spec = CATALOG_BY_KEY.get(key)
    is_catalog_field = spec is not None
    if spec is None:
        label = str(raw.get("label") or raw.get("alias") or "Custom field").strip()
        alias = safe_alias(str(raw.get("alias") or label))
        expression = str(raw.get("expression") or "").strip()
        field_type = str(raw.get("field_type") or raw.get("type") or "numeric")
        spec = FieldSpec(
            key=f"custom_{alias}",
            label=label,
            alias=alias,
            expression=expression,
            field_type="text" if field_type == "text" else "numeric",
            definition=str(raw.get("definition") or "Custom Bloomberg/BQL field."),
            default_input=bool(raw.get("input", True)),
            default_output=bool(raw.get("output", True)),
            format_type="text" if field_type == "text" else "number",
        )
    supports_window = field_supports_window(spec.expression)
    n_quarters = int(raw.get("n_quarters") or DEFAULT_STABILITY_QUARTERS)
    expression = (
        spec.expression_for(currency=currency, n_quarters=n_quarters)
        if is_catalog_field or field_uses_placeholders(spec.expression)
        else str(raw.get("expression") or spec.expression)
    )
    return {
        "key": spec.key,
        "label": str(raw.get("label") or spec.label),
        "alias": safe_alias(str(raw.get("alias") or spec.alias)),
        "expression": expression,
        "field_type": spec.field_type,
        "definition": spec.definition,
        "input": bool(raw.get("input", spec.default_input)),
        "output": bool(raw.get("output", spec.default_output)),
        "min": str(raw.get("min", spec.default_min)),
        "max": str(raw.get("max", spec.default_max)),
        "text": str(raw.get("text", spec.default_text)),
        "n_quarters": n_quarters,
        "supports_window": supports_window,
        "format_type": spec.format_type,
    }


def safe_alias(value: str) -> str:
    alias = re.sub(r"[^A-Za-z0-9_]+", "_", value.strip().lower()).strip("_")
    if not alias:
        alias = "field"
    if alias[0].isdigit():
        alias = f"field_{alias}"
    return alias


def build_query(config: dict) -> str:
    geography = config.get("geography", "western_europe")
    currency = config.get("currency", "EUR")
    countries = list(COUNTRY_GROUPS.get(geography, WESTERN_EUROPE_COUNTRIES).values())
    country_expr = "[" + ",".join(bql_quote(code) for code in countries) + "]"
    fields = [normalize_field(item, currency) for item in config.get("fields", [])]

    output_fields = [field for field in fields if field["output"]]
    selected_aliases = {field["alias"] for field in output_fields}
    get_fields = []
    seen = set()
    for identity in IDENTITY_FIELDS:
        if identity.alias not in selected_aliases and identity.expression not in seen:
            get_fields.append(identity.expression)
            seen.add(identity.expression)
    for field in output_fields:
        expression = field["expression"]
        if expression and expression not in seen:
            get_fields.append(expression)
            seen.add(expression)

    conditions = [
        f"cntry_of_domicile in {country_expr}",
        "security_typ == 'Common Stock'",
    ]
    for field in fields:
        if not field["input"] or not field["expression"]:
            continue
        if field["field_type"] == "numeric":
            add_numeric_conditions(conditions, field["expression"], field["min"], field["max"])
        else:
            values = parse_text_values(field["text"])
            if len(values) == 1:
                conditions.append(f"{field['expression']} == {bql_quote(values[0])}")
            elif values:
                conditions.append(f"{field['expression']} in [" + ",".join(bql_quote(value) for value in values) + "]")

    get_text = ",\n".join(f"    {expression}" for expression in get_fields)
    condition_text = "\n        and ".join(conditions)
    return f"""get(
{get_text}
)
for(
    filter(
        equitiesUniv([ACTIVE, PRIMARY]),
        {condition_text}
    )
)"""


def build_workbook(config: dict) -> bytes:
    currency = config.get("currency", "EUR")
    fields = [normalize_field(item, currency) for item in config.get("fields", [])]
    config = {**config, "fields": fields, "generated_at": datetime.now().isoformat(timespec="seconds")}
    query = build_query(config)
    output_columns = output_columns_for(fields)
    buffer = BytesIO()
    workbook = xlsxwriter.Workbook(buffer, {"in_memory": True})
    header_format = workbook.add_format({"bold": True, "font_color": "white", "bg_color": "#1F4E78", "valign": "top"})
    title_format = workbook.add_format({"bold": True, "font_color": "#1F4E78", "font_size": 16})
    bold_blue = workbook.add_format({"bold": True, "font_color": "#1F4E78", "valign": "top"})
    wrap = workbook.add_format({"text_wrap": True, "valign": "top"})

    setup = workbook.add_worksheet(SETUP_SHEET)
    setup.write("A1", "Setup Summary", title_format)
    setup.write_row("A3", ["Generated at", config["generated_at"]])
    setup.write_row("A4", ["Geography", config.get("geography", "western_europe")])
    setup.write_row("A5", ["Currency", config.get("currency", "EUR")])
    setup.write_row("A6", ["Raw output", "Bloomberg BQL output spills on Raw Data from A1. Keep Raw Data columns A:AZ clear."])
    setup.write_row("A7", ["Sort field", config.get("sort_field") or default_sort_field(output_columns)])
    setup.write_row("A8", ["Sort direction", config.get("sort_direction", "Ascending")])
    setup.data_validation("B7", {"validate": "list", "source": [column["label"] for column in output_columns]})
    setup.data_validation("B8", {"validate": "list", "source": ["Ascending", "Descending"]})
    setup.write("A10", "Generated BQL Query", header_format)
    setup.write("B10", query, wrap)
    setup.write("A13", "Selected Field Summary", title_format)
    setup.write_row("A15", ["Field", "Alias", "Input", "Output", "Min", "Max", "Text filter", "Window", "BQL expression", "Definition"], header_format)
    for row_index, field in enumerate(config["fields"], start=16):
        setup.write_row(
            row_index - 1,
            0,
            [
                field["label"],
                field["alias"],
                "Y" if field["input"] else "",
                "Y" if field["output"] else "",
                field["min"],
                field["max"],
                field["text"],
                field["n_quarters"],
                field["expression"],
                field["definition"],
            ],
            wrap,
        )
    set_xlsxwriter_widths(setup, [28, 22, 10, 10, 14, 14, 22, 12, 42, 70])
    setup.freeze_panes(15, 0)

    raw_data = workbook.add_worksheet(RAW_DATA_SHEET)
    raw_data.write_dynamic_array_formula("A1", f"=BQL.QUERY({SETUP_QUERY_CELL})", bold_blue)
    raw_data.set_column(0, 51, 18)
    raw_data.freeze_panes(1, 0)

    clean_view = workbook.add_worksheet("Clean View")
    for col_index, column in enumerate(output_columns):
        clean_view.write(0, col_index, column["label"], header_format)
        clean_view.set_column(col_index, col_index, max(14, min(34, len(column["label"]) + 4)))
    if output_columns:
        clean_view.write_dynamic_array_formula(1, 0, 1, 0, clean_view_spill_formula(output_columns, ooxml=True))
    clean_view.freeze_panes(1, 0)

    workbook.close()
    return buffer.getvalue()


def build_clean_workbook(config: dict, raw_text: str) -> bytes:
    currency = config.get("currency", "EUR")
    fields = [normalize_field(item, currency) for item in config.get("fields", [])]
    columns = output_columns_for(fields)
    clean_rows = clean_raw_bql_rows(raw_text, columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "Clean View"
    for col_index, column in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=column["label"])
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row_index, row in enumerate(clean_rows, start=2):
        for col_index, column in enumerate(columns, start=1):
            ws.cell(row=row_index, column=col_index, value=row.get(column["label"], ""))
    last_col = get_column_letter(max(1, len(columns)))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, len(clean_rows) + 1)}"
    ws.freeze_panes = "A2"
    for index, column in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(index)].width = max(14, min(34, len(column["label"]) + 4))

    raw_ws = wb.create_sheet("Raw Paste")
    for row in parse_tsv(raw_text):
        raw_ws.append(row)
    raw_ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def output_aliases_for(fields: list[dict]) -> list[str]:
    output_columns = output_columns_for(fields)
    return [column["label"] for column in output_columns]


def default_sort_field(columns: list[dict[str, str]]) -> str:
    labels = [column["label"] for column in columns]
    for preferred in ["margin_stability_pp", "market_cap", "name", "id"]:
        if preferred in labels:
            return preferred
    return labels[0] if labels else "id"


def output_columns_for(fields: list[dict]) -> list[dict[str, str]]:
    columns = [
        {"label": "id", "raw_header": "ID"},
        {"label": "name", "raw_header": "name()"},
        {"label": "cntry_of_domicile", "raw_header": "cntry_of_domicile()"},
        {"label": "crncy", "raw_header": "crncy()"},
        {"label": "sector", "raw_header": "classification_name(BICS,1)"},
        {"label": "industry_group", "raw_header": "classification_name(BICS,2)"},
    ]
    labels = {column["label"] for column in columns}
    for field in fields:
        if field["output"] and field["alias"] not in labels:
            columns.append({"label": field["alias"], "raw_header": bql_result_header(field["expression"])})
            labels.add(field["alias"])
    return columns


def parse_tsv(raw_text: str) -> list[list[str]]:
    text = raw_text.replace("\r\n", "\n").replace("\r", "\n").strip("\n")
    if not text.strip():
        return []
    return [row for row in csv.reader(text.splitlines(), dialect="excel-tab")]


def normalize_raw_header(value: str) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def clean_raw_bql_rows(raw_text: str, columns: list[dict[str, str]]) -> list[dict[str, str]]:
    rows = parse_tsv(raw_text)
    if not rows:
        return []
    headers = [normalize_raw_header(value) for value in rows[0]]
    header_index = {header: index for index, header in enumerate(headers) if header}
    wanted = [(column["label"], normalize_raw_header(column["raw_header"])) for column in columns]

    by_id: dict[str, dict[str, str]] = {}
    order: list[str] = []
    for raw_row in rows[1:]:
        if not raw_row:
            continue
        ticker = raw_row[0].strip() if len(raw_row) > 0 else ""
        if not ticker:
            continue
        if ticker not in by_id:
            by_id[ticker] = {"id": ticker}
            order.append(ticker)
        out = by_id[ticker]
        for label, raw_header in wanted:
            if label == "id":
                continue
            index = header_index.get(raw_header)
            if index is None or index >= len(raw_row):
                continue
            value = raw_row[index].strip()
            if value and not out.get(label):
                out[label] = value
    return [by_id[ticker] for ticker in order]


def office_script_code(columns: list[dict[str, str]]) -> str:
    columns_json = json.dumps(columns, indent=2)
    return f"""function main(workbook: ExcelScript.Workbook) {{
  const columns: {{ label: string; raw_header: string }}[] = {columns_json};
  const raw = workbook.getWorksheet("Raw Data");
  const clean = workbook.getWorksheet("Clean View");
  const control = workbook.getWorksheet("Setup");
  const used = raw.getUsedRange();
  if (!used) {{
    throw new Error("Raw Data is empty. Refresh Bloomberg first.");
  }}

  const values = used.getTexts();
  if (values.length < 2) {{
    throw new Error("Raw Data has no data rows.");
  }}

  const normalize = (value: string) => String(value || "").replace(/\\s+/g, "");
  const headers = values[0].map(normalize);
  const headerIndex = new Map<string, number>();
  headers.forEach((header, index) => {{
    if (header) headerIndex.set(header, index);
  }});

  const order: string[] = [];
  const byId = new Map<string, Record<string, string>>();
  for (let r = 1; r < values.length; r++) {{
    const row = values[r];
    const id = String(row[0] || "").trim();
    if (!id) continue;
    if (!byId.has(id)) {{
      byId.set(id, {{ id }});
      order.push(id);
    }}
    const output = byId.get(id)!;
    for (const column of columns) {{
      if (column.label === "id") continue;
      const idx = headerIndex.get(normalize(column.raw_header));
      if (idx === undefined) continue;
      const value = String(row[idx] || "").trim();
      if (value && !output[column.label]) output[column.label] = value;
    }}
  }}

  const sortField = String(control.getRange("B7").getText() || "id");
  const sortDirection = String(control.getRange("B8").getText() || "Ascending");
  const sortSign = sortDirection === "Descending" ? -1 : 1;
  const rows = order.map(id => byId.get(id)!);
  rows.sort((a, b) => {{
    const av = a[sortField] || "";
    const bv = b[sortField] || "";
    const an = Number(String(av).replace(",", "."));
    const bn = Number(String(bv).replace(",", "."));
    if (!Number.isNaN(an) && !Number.isNaN(bn)) return (an - bn) * sortSign;
    return String(av).localeCompare(String(bv)) * sortSign;
  }});

  clean.getUsedRange()?.clear(ExcelScript.ClearApplyTo.all);
  const output = [
    columns.map(column => column.label),
    ...rows.map(row => columns.map(column => row[column.label] || ""))
  ];
  const target = clean.getRangeByIndexes(0, 0, output.length, columns.length);
  target.setValues(output);

  clean.getTables().forEach(table => table.delete());
  const table = clean.addTable(target, true);
  table.setName("AdaptiveCleanView");
  clean.getFreezePanes().freezeRows(1);
  target.getFormat().autofitColumns();
}}"""


def bql_result_header(expression: str) -> str:
    return re.sub(r"\s+", "", expression)


def excel_quote(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def clean_view_rows(columns: list[dict[str, str]]) -> list[list[str]]:
    rows = [[column["label"] for column in columns]]
    if not columns:
        return rows
    rows.append([clean_view_spill_formula(columns, ooxml=True)])
    return rows


def excel_modern_function(name: str, *, ooxml: bool) -> str:
    if not ooxml:
        return name
    if name == "FILTER":
        return "_xlfn._xlws.FILTER"
    return f"_xlfn.{name}"


def clean_view_spill_formula(columns: list[dict[str, str]], *, ooxml: bool = False) -> str:
    let_fn = excel_modern_function("LET", ooxml=ooxml)
    filter_fn = excel_modern_function("FILTER", ooxml=ooxml)
    sort_fn = excel_modern_function("SORT", ooxml=ooxml)
    sortby_fn = excel_modern_function("SORTBY", ooxml=ooxml)
    unique_fn = excel_modern_function("UNIQUE", ooxml=ooxml)
    xlookup_fn = excel_modern_function("XLOOKUP", ooxml=ooxml)
    hstack_fn = excel_modern_function("HSTACK", ooxml=ooxml)
    var = (lambda name: f"_xlpm.{name}") if ooxml else (lambda name: name)

    definitions = [
        f"{var('rawIds')},'Raw Data'!$A$2:$A$5000",
        f"{var('rawData')},'Raw Data'!$A$2:$AZ$5000",
        f"{var('headers')},'Raw Data'!$A$1:$AZ$1",
        f'{var("ids")},{sort_fn}({unique_fn}({filter_fn}({var("rawIds")},{var("rawIds")}<>"")))',
    ]
    values = [var("ids")]
    for index, column in enumerate(columns[1:], start=1):
        col_name = f"c{index}"
        value_name = f"v{index}"
        definitions.append(f"{var(col_name)},INDEX({var('rawData')},0,MATCH({excel_quote(column['raw_header'])},{var('headers')},0))")
        definitions.append(
            f'{var(value_name)},{xlookup_fn}({var("ids")},'
            f'{filter_fn}({var("rawIds")},{var(col_name)}<>""),'
            f'{filter_fn}({var(col_name)},{var(col_name)}<>""),"")'
        )
        values.append(var(value_name))
    hstack_args = ",".join(values)
    header_array = "{" + ",".join(excel_quote(column["label"]) for column in columns) + "}"
    return (
        f"={let_fn}("
        + ",".join(definitions)
        + ","
        + f"{var('table')},{hstack_fn}({hstack_args}),"
        + f"{var('sortField')},{SETUP_SORT_FIELD_CELL},"
        + f"{var('sortDir')},IF({SETUP_SORT_DIRECTION_CELL}=\"Descending\",-1,1),"
        + f"{var('sortCol')},IFERROR(MATCH({var('sortField')},{header_array},0),1),"
        + f"{sortby_fn}({var('table')},INDEX({var('table')},0,{var('sortCol')}),{var('sortDir')})"
        ")"
    )


def build_xml_workbook(config: dict) -> bytes:
    """Build an Excel-readable XML Spreadsheet 2003 workbook.

    This is a single XML file with an .xls extension, so it is not a zipped
    Office Open XML package like .xlsx.
    """
    currency = config.get("currency", "EUR")
    fields = [normalize_field(item, currency) for item in config.get("fields", [])]
    config = {**config, "fields": fields, "generated_at": datetime.now().isoformat(timespec="seconds")}
    query = build_query(config)

    output_columns = output_columns_for(fields)

    sheets = [
        (
            "Control Panel",
            [
                ["Adaptive Screener Excel Generator"],
                [],
                ["Generated at", config["generated_at"]],
                ["Geography", config.get("geography", "western_europe")],
                ["Currency", config.get("currency", "EUR")],
                ["Notes", "This workbook stores the selected setup. Raw Bloomberg output lands on Raw BQL; Clean View turns it into one row per company."],
                [],
                ["Sort field", config.get("sort_field") or default_sort_field(output_columns)],
                ["Sort direction", config.get("sort_direction", "Ascending")],
                [],
                ["Field", "Alias", "Input", "Output", "Min", "Max", "Text filter", "Window", "BQL expression", "Definition"],
                *[
                    [
                        field["label"],
                        field["alias"],
                        "Y" if field["input"] else "",
                        "Y" if field["output"] else "",
                        field["min"],
                        field["max"],
                        field["text"],
                        field["n_quarters"],
                        field["expression"],
                        field["definition"],
                    ]
                    for field in fields
                ],
            ],
        ),
        (
            "BQL Query",
            [
                ["Generated BQL Query"],
                [query],
                [],
                ["The executable Bloomberg Excel formula is on the Raw BQL sheet at A1 so the returned table has room to spill."],
            ],
        ),
        ("Raw BQL", [["=BQL.QUERY('BQL Query'!A2)"]]),
        ("Clean View", [[column["label"] for column in output_columns]]),
        (
            "Field Catalog",
            [
                ["Key", "Label", "Alias", "Expression", "Type", "Default input", "Default output", "Default min", "Default max", "Definition"],
                *[
                    [
                        field.key,
                        field.label,
                        field.alias,
                        field.expression,
                        field.field_type,
                        "Y" if field.default_input else "",
                        "Y" if field.default_output else "",
                        field.default_min,
                        field.default_max,
                        field.definition,
                    ]
                    for field in FIELD_CATALOG
                ],
            ],
        ),
        ("Presets", [["Preset/config JSON"], [json.dumps({"version": 1, "config": config, "query": query}, indent=2, sort_keys=True)]]),
    ]

    workbook = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<?mso-application progid="Excel.Sheet"?>',
        '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"',
        ' xmlns:o="urn:schemas-microsoft-com:office:office"',
        ' xmlns:x="urn:schemas-microsoft-com:office:excel"',
        ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"',
        ' xmlns:html="http://www.w3.org/TR/REC-html40">',
        "<Styles>",
        '<Style ss:ID="Header"><Font ss:Bold="1" ss:Color="#FFFFFF"/><Interior ss:Color="#1F4E78" ss:Pattern="Solid"/></Style>',
        '<Style ss:ID="Wrap"><Alignment ss:Vertical="Top" ss:WrapText="1"/></Style>',
        "</Styles>",
    ]
    for sheet_name, rows in sheets:
        workbook.append(f'<Worksheet ss:Name="{xml_escape(sheet_name)}"><Table>')
        for row_index, row in enumerate(rows, start=1):
            style = ' ss:StyleID="Header"' if row_index == 1 or (sheet_name == "Control Panel" and row_index == 8) else ' ss:StyleID="Wrap"'
            workbook.append("<Row>")
            for value in row:
                workbook.append(xml_cell(value, style))
            workbook.append("</Row>")
        workbook.append("</Table>")
        if sheet_name == "Clean View":
            workbook.append("<WorksheetOptions xmlns=\"urn:schemas-microsoft-com:office:excel\"><FreezePanes/><FrozenNoSplit/><SplitHorizontal>1</SplitHorizontal><TopRowBottomPane>1</TopRowBottomPane></WorksheetOptions>")
        workbook.append("</Worksheet>")
    workbook.append("</Workbook>")
    return "\n".join(workbook).encode("utf-8")


def xml_cell(value: object, style: str) -> str:
    if isinstance(value, str) and value.startswith("="):
        return f'<Cell{style} ss:Formula="{xml_escape(value)}"><Data ss:Type="String"></Data></Cell>'
    cell_type = "Number" if isinstance(value, (int, float)) and not isinstance(value, bool) else "String"
    return f"<Cell{style}><Data ss:Type=\"{cell_type}\">{xml_escape(value)}</Data></Cell>"


def xml_escape(value: object) -> str:
    return html.escape(str(value), quote=True)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill


def set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def set_xlsxwriter_widths(ws, widths: list[int]) -> None:
    for index, width in enumerate(widths):
        ws.set_column(index, index, width)


def write_control_sheet(ws, config: dict) -> None:
    ws.append(["Adaptive Screener Excel Generator"])
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    ws.append([])
    ws.append(["Generated at", config["generated_at"]])
    ws.append(["Geography", config.get("geography", "western_europe")])
    ws.append(["Currency", config.get("currency", "EUR")])
    ws.append(["Notes", "This workbook stores the selected setup. Raw Bloomberg output lands on Raw BQL; Clean View turns it into one row per company."])
    ws.append([])
    ws.append(["Field", "Alias", "Input", "Output", "Min", "Max", "Text filter", "Window", "BQL expression", "Definition"])
    style_header(ws, 8)
    for field in config["fields"]:
        ws.append(
            [
                field["label"],
                field["alias"],
                "Y" if field["input"] else "",
                "Y" if field["output"] else "",
                field["min"],
                field["max"],
                field["text"],
                field["n_quarters"],
                field["expression"],
                field["definition"],
            ]
        )
    set_widths(ws, {"A": 28, "B": 22, "C": 10, "D": 10, "E": 14, "F": 14, "G": 22, "H": 12, "I": 42, "J": 60})


def write_query_sheet(ws, query: str) -> None:
    ws.append(["Generated BQL Query"])
    style_header(ws)
    ws.append([query])
    ws.append([])
    ws.append(["The executable Bloomberg Excel formula is on the Raw BQL sheet at A1 so the returned table has room to spill."])
    set_widths(ws, {"A": 120})


def write_raw_results_sheet(ws) -> None:
    ws["A1"] = "=BQL.QUERY('BQL Query'!A2)"
    ws["A1"].font = Font(bold=True, color="1F4E78")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24


def write_clean_view_sheet(ws, fields: list[dict]) -> None:
    output_columns = output_columns_for(fields)
    for col_index, column in enumerate(output_columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=column["label"])
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    for index, column in enumerate(output_columns, start=1):
        alias = column["label"]
        ws.column_dimensions[get_column_letter(index)].width = max(14, min(32, len(alias) + 4))


def write_clean_formula_sheet(ws, fields: list[dict]) -> None:
    output_columns = output_columns_for(fields)
    ws.append(["Formula for Clean View!A2"])
    style_header(ws)
    formula_cell = ws["A2"]
    formula_cell.value = clean_view_spill_formula(output_columns, ooxml=False)
    formula_cell.data_type = "s"
    formula_cell.number_format = "@"
    formula_cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws["A4"] = "Copy A2, go to Clean View!A2, paste, then calculate. This keeps the generated workbook repair-safe."
    set_widths(ws, {"A": 140})


def write_catalog_sheet(ws) -> None:
    ws.append(["Key", "Label", "Alias", "Expression", "Type", "Default input", "Default output", "Default min", "Default max", "Definition"])
    style_header(ws)
    for field in FIELD_CATALOG:
        ws.append(
            [
                field.key,
                field.label,
                field.alias,
                field.expression,
                field.field_type,
                "Y" if field.default_input else "",
                "Y" if field.default_output else "",
                field.default_min,
                field.default_max,
                field.definition,
            ]
        )
    set_widths(ws, {"A": 28, "B": 34, "C": 24, "D": 42, "E": 12, "F": 14, "G": 14, "H": 14, "I": 14, "J": 70})


def write_presets_sheet(ws, config: dict, query: str) -> None:
    ws.append(["Preset/config JSON"])
    style_header(ws)
    ws.append([json.dumps({"version": 1, "config": config, "query": query}, indent=2, sort_keys=True)])
    ws["A2"].alignment = Alignment(vertical="top", wrap_text=True)
    set_widths(ws, {"A": 120})


def load_user_templates() -> list[dict]:
    if not USER_TEMPLATE_PATH.exists():
        return []
    try:
        data = json.loads(USER_TEMPLATE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    templates = data.get("templates", data) if isinstance(data, dict) else data
    if not isinstance(templates, list):
        return []
    return [{**template, "user_template": True} for template in templates if isinstance(template, dict)]


def save_user_templates(templates: list[dict]) -> None:
    clean_templates = []
    for template in templates:
        if not isinstance(template, dict):
            continue
        clean_templates.append({**template, "user_template": True})
    payload = {
        "version": 1,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "templates": clean_templates,
    }
    USER_TEMPLATE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def upsert_user_template(template: dict) -> dict:
    label = str(template.get("label") or "").strip()
    if not label:
        raise ValueError("Template name is required.")
    key = str(template.get("key") or "").strip()
    if not key or not key.startswith("user_"):
        key = f"user_{safe_alias(label)}_{int(datetime.now().timestamp())}"
    saved = {**template, "key": key, "label": label, "user_template": True}
    templates = load_user_templates()
    replaced = False
    next_templates = []
    for current in templates:
        if current.get("key") == key:
            next_templates.append(saved)
            replaced = True
        else:
            next_templates.append(current)
    if not replaced:
        next_templates.append(saved)
    save_user_templates(next_templates)
    return saved


def delete_user_template(key: str) -> list[dict]:
    templates = load_user_templates()
    next_templates = [template for template in templates if template.get("key") != key]
    if len(next_templates) == len(templates):
        raise ValueError("Saved template was not found.")
    save_user_templates(next_templates)
    return next_templates


def catalog_payload() -> dict:
    user_templates = load_user_templates()
    return {
        "fields": [
            {**asdict(field), "supports_window": field_supports_window(field.expression)}
            for field in FIELD_CATALOG
        ],
        "default_field_keys": DEFAULT_FIELD_KEYS,
        "templates": FIELD_TEMPLATES,
        "user_templates": user_templates,
        "template_path": str(USER_TEMPLATE_PATH),
        "geographies": [
            {"key": key, "label": label, "countries": countries}
            for key, label, countries in [
                ("western_europe", "Western Europe", WESTERN_EUROPE_COUNTRIES),
                ("north_america", "North America", NORTH_AMERICA_COUNTRIES),
            ]
        ],
    }


HTML_PAGE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Adaptive Screener Excel Generator</title>
<style>
:root{font-family:Arial,Helvetica,sans-serif;color:#18212f;background:#f6f7f9}
body{margin:0}
main{max-width:1180px;margin:0 auto;padding:18px}
h1{font-size:24px;margin:0 0 4px}
h2{font-size:15px;margin:0 0 10px;color:#1f4e78}
.top{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap;margin-bottom:14px}
.panel{background:white;border:1px solid #d9dee7;border-radius:8px;padding:14px;margin-bottom:12px}
.row{display:flex;gap:10px;flex-wrap:wrap;align-items:flex-end}
label{font-size:12px;color:#536070;display:flex;flex-direction:column;gap:4px}
select,input,textarea,button{font:inherit}
select,input,textarea{border:1px solid #b9c2cf;border-radius:6px;padding:7px;background:white}
button{border:1px solid #1f4e78;background:#1f4e78;color:white;border-radius:6px;padding:8px 11px;cursor:pointer}
button.secondary{background:white;color:#1f4e78}
button.danger{background:white;color:#a01818;border-color:#a01818}
.template-bar{align-items:center}
.template-select{min-width:260px}
.template-description{flex:1;min-width:220px}
.template-manage{margin-top:10px}
.template-manage summary{cursor:pointer;color:#1f4e78;font-size:12px;font-weight:700}
.template-manage .row{margin-top:10px}
.search{width:min(520px,100%)}
.matches{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:8px;margin-top:10px}
.match{border:1px solid #d8dde5;border-radius:6px;background:#fbfcfd;color:#18212f;text-align:left}
.field-card{border:1px solid #d8dde5;border-radius:8px;padding:10px;margin:8px 0;background:#fcfcfd}
.field-grid{display:grid;grid-template-columns:minmax(220px,1.4fr) 80px 90px 120px 120px 120px 1fr 82px;gap:8px;align-items:center}
.field-title{font-weight:700}.field-expression{font-size:12px;color:#536070;margin-top:2px}.definition{display:none;color:#536070;font-size:12px;margin-top:8px}
.tiny{font-size:12px;color:#536070}.status{min-height:22px;color:#1f4e78}.hidden{display:none}
@media(max-width:860px){.field-grid{grid-template-columns:1fr 80px 90px}.wide{grid-column:1/-1}}
</style>
</head>
<body>
<main>
  <div class="top">
    <div>
      <h1>Adaptive Screener Excel Generator</h1>
      <div class="tiny">Configure locally, then download an Excel workbook that stores the setup and generated BQL query.</div>
    </div>
  </div>

  <section class="panel">
    <h2>Template</h2>
    <div class="row template-bar">
      <label class="template-select">Template<select id="template"></select></label>
      <button id="applyTemplate" class="secondary">Load</button>
      <button id="updateTemplate" class="secondary">Update</button>
      <button id="deleteTemplate" class="danger">Delete</button>
      <span id="templateDescription" class="tiny template-description"></span>
    </div>
    <details class="template-manage">
      <summary>Create template</summary>
      <div class="row">
        <label>Name<input id="templateName" placeholder="Template name"></label>
        <button id="saveTemplate" class="secondary">Create</button>
      </div>
    </details>
  </section>

  <section class="panel">
    <h2>Universe</h2>
    <div class="row">
      <label>Geography<select id="geography"></select></label>
      <label>Currency<select id="currency"><option>EUR</option><option>GBP</option><option>USD</option><option>SEK</option><option>DKK</option><option>NOK</option><option>CHF</option></select></label>
    </div>
  </section>

  <section class="panel">
    <h2>Fields</h2>
    <div class="row">
      <label class="search">Search labels<input id="search" placeholder="Search labels"></label>
      <button id="add" class="secondary">Add selected</button>
    </div>
    <div id="matches" class="matches"></div>
    <details style="margin-top:12px">
      <summary>Custom Bloomberg/BQL field</summary>
      <div class="row" style="margin-top:10px">
        <label>Label<input id="customLabel" placeholder="Custom label"></label>
        <label>Alias<input id="customAlias" placeholder="custom_alias"></label>
        <label>Type<select id="customType"><option value="numeric">Numeric</option><option value="text">Text</option></select></label>
      </div>
      <div class="row" style="margin-top:8px">
        <label style="flex:1">BQL expression<input id="customExpression" placeholder="px_to_book_ratio"></label>
        <label style="flex:1">Definition<input id="customDefinition" placeholder="What this field means"></label>
        <button id="addCustom" class="secondary">Add custom</button>
      </div>
    </details>
    <div id="selected"></div>
  </section>

  <section class="panel">
    <div class="row">
      <button id="download">Download .xlsx</button>
      <span id="status" class="status"></span>
    </div>
  </section>
</main>
<script>
let catalog = [], defaults = [], builtinTemplates = [], userTemplates = [], templates = [], selected = [], activeMatch = null;
const $ = id => document.getElementById(id);
const escapeHtml = s => String(s ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
const safeAlias = s => String(s || 'field').trim().toLowerCase().replace(/[^a-z0-9_]+/g,'_').replace(/^_+|_+$/g,'') || 'field';

async function init(){
  const data = await fetch('/catalog').then(r => r.json());
  catalog = data.fields; defaults = data.default_field_keys; builtinTemplates = data.templates || [];
  userTemplates = data.user_templates || [];
  templates = [...userTemplates];
  $('geography').innerHTML = data.geographies.map(g => `<option value="${g.key}">${g.label}</option>`).join('');
  const initialTemplateKey = templates[0]?.key || '';
  renderTemplateSelect(initialTemplateKey);
  applyTemplate(initialTemplateKey);
}

function selectedFromSpec(spec){
  const base = catalog.find(f => f.key === spec.key);
  if(!base && !spec.expression) return null;
  if(!base){
    return {
      key: spec.key || 'custom_' + Date.now(),
      label: spec.label || spec.alias || 'Custom field',
      alias: safeAlias(spec.alias || spec.label || 'custom_field'),
      expression: spec.expression || '',
      field_type: spec.field_type === 'text' ? 'text' : 'numeric',
      definition: spec.definition || 'Custom Bloomberg/BQL field.',
      input: spec.input ?? true,
      output: spec.output ?? true,
      min: spec.min ?? '',
      max: spec.max ?? '',
      text: spec.text ?? '',
      n_quarters: Number(spec.n_quarters || 12),
      supports_window: Boolean(spec.supports_window || String(spec.expression || '').includes('{fpo_start}') || String(spec.expression || '').includes('{n_quarters}')),
    };
  }
  return {
    ...base,
    input: spec.input ?? base.default_input,
    output: spec.output ?? base.default_output,
    min: spec.min ?? base.default_min ?? '',
    max: spec.max ?? base.default_max ?? '',
    text: spec.text ?? base.default_text ?? '',
    n_quarters: Number(spec.n_quarters || 12),
  };
}

function setUserTemplates(savedTemplates){
  userTemplates = (savedTemplates || []).map(t => ({...t, user_template:true}));
  templates = [...userTemplates];
}

function renderTemplateSelect(selectedKey){
  const customOptions = userTemplates.map(t => `<option value="${escapeHtml(t.key)}">${escapeHtml(t.label)}</option>`).join('');
  $('template').innerHTML = customOptions || '<option value="">No saved templates</option>';
  if(selectedKey && templates.some(t => t.key === selectedKey)) $('template').value = selectedKey;
}

function templateByKey(key){
  return templates.find(t => t.key === key);
}

function currentTemplatePayload(label, key){
  return {
    key,
    label,
    description: `Saved ${new Date().toLocaleString()}`,
    geography: $('geography').value,
    currency: $('currency').value,
    sort_field: defaultSortField(),
    sort_direction: 'Ascending',
    user_template: true,
    fields: selected.map(f => ({
      key: f.key,
      label: f.label,
      alias: f.alias,
      expression: f.expression,
      field_type: f.field_type,
      definition: f.definition,
      input: Boolean(f.input),
      output: Boolean(f.output),
      min: f.min || '',
      max: f.max || '',
      text: f.text || '',
      n_quarters: Number(f.n_quarters || 12),
      supports_window: Boolean(f.supports_window),
    })),
  };
}

async function saveTemplate({overwrite=false} = {}){
  const selectedTemplate = templateByKey($('template').value);
  const selectedIsUser = Boolean(selectedTemplate?.user_template);
  if(overwrite && !selectedIsUser){
    $('status').textContent = 'Choose a saved template to update.';
    return;
  }
  const label = $('templateName').value.trim() || (overwrite && selectedIsUser ? selectedTemplate.label : '');
  if(!label){
    $('status').textContent = 'Give the template a name first.';
    return;
  }
  const key = overwrite && selectedIsUser ? selectedTemplate.key : 'user_' + safeAlias(label) + '_' + Date.now();
  const payload = currentTemplatePayload(label, key);
  const r = await fetch('/template-save', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(payload)});
  const text = await r.text();
  if(!r.ok){ $('status').textContent = text; return; }
  const result = JSON.parse(text);
  setUserTemplates(result.user_templates);
  renderTemplateSelect(result.template.key);
  $('templateDescription').textContent = result.template.description;
  $('templateName').value = result.template.label;
  $('status').textContent = overwrite && selectedIsUser ? 'Template updated.' : 'Template created.';
}

async function deleteTemplate(){
  const selectedTemplate = templateByKey($('template').value);
  if(!selectedTemplate?.user_template){
    $('status').textContent = 'Choose a saved template to delete.';
    return;
  }
  const r = await fetch('/template-delete', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify({key:selectedTemplate.key})});
  const text = await r.text();
  if(!r.ok){ $('status').textContent = text; return; }
  const result = JSON.parse(text);
  setUserTemplates(result.user_templates);
  const nextKey = templates[0]?.key || '';
  renderTemplateSelect(nextKey);
  applyTemplate(nextKey);
  $('status').textContent = 'Template deleted.';
}

function applyTemplate(key){
  const template = templateByKey(key);
  if(!template){
    selected = defaults.map(key => selectedFromSpec({key})).filter(Boolean);
    $('templateDescription').textContent = 'No saved template selected. Using the default field setup.';
    $('templateName').value = '';
    renderMatches(); renderSelected(); refreshQuery();
    return;
  }
  $('template').value = template.key;
  $('templateDescription').textContent = template.description || '';
  $('templateName').value = template.user_template ? template.label : '';
  if(template.geography) $('geography').value = template.geography;
  if(template.currency) $('currency').value = template.currency;
  selected = (template.fields || []).map(selectedFromSpec).filter(Boolean);
  renderMatches(); renderSelected();
  refreshQuery();
}

function score(field, q){
  q = q.trim().toLowerCase();
  if(!q) return 0;
  const hay = [field.label, field.alias, field.expression, field.definition].join(' ').toLowerCase();
  if(hay.includes(q)) return 2 + q.length / Math.max(hay.length, 1);
  let hits = 0; for(const part of q.split(/\s+/)){ if(hay.includes(part)) hits++; }
  return hits / Math.max(q.split(/\s+/).length, 1);
}

function renderMatches(){
  const q = $('search').value;
  const matches = [...catalog].filter(f => !selected.some(s => s.key === f.key)).sort((a,b)=>score(b,q)-score(a,q)).filter(f => q.trim() && score(f,q)>0).slice(0,5);
  activeMatch = matches[0]?.key || null;
  $('matches').innerHTML = matches.map(f => `<button class="match" data-key="${f.key}"><b>${escapeHtml(f.label)}</b><br><span class="tiny">${escapeHtml(f.expression)}</span></button>`).join('');
  document.querySelectorAll('.match').forEach(btn => btn.onclick = () => { activeMatch = btn.dataset.key; document.querySelectorAll('.match').forEach(b => b.style.outline=''); btn.style.outline='2px solid #1f4e78'; });
}

function addField(key){
  const f = catalog.find(x => x.key === key);
  if(!f || selected.some(s => s.key === key)) return;
  selected.push({...f, input:f.default_input, output:f.default_output, min:f.default_min || '', max:f.default_max || '', text:f.default_text || '', n_quarters:12});
  renderMatches(); renderSelected(); refreshQuery();
}

function outputAliases(){
  const aliases = ['id','name','cntry_of_domicile','crncy','sector','industry_group'];
  selected.forEach(f => { if(f.output && !aliases.includes(f.alias)) aliases.push(f.alias); });
  return aliases;
}

function defaultSortField(){
  const aliases = outputAliases();
  if(aliases.includes('margin_stability_pp')) return 'margin_stability_pp';
  if(aliases.includes('market_cap')) return 'market_cap';
  return aliases[0] || 'id';
}

function renderSelected(){
  $('selected').innerHTML = selected.map((f,i) => `
    <div class="field-card">
      <div class="field-grid">
        <div class="wide"><div class="field-title">${escapeHtml(f.label)}</div><div class="field-expression">${escapeHtml(f.expression)}</div></div>
        <label><input type="checkbox" data-i="${i}" data-k="input" ${f.input?'checked':''}> Input</label>
        <label><input type="checkbox" data-i="${i}" data-k="output" ${f.output?'checked':''}> Output</label>
        ${f.field_type === 'numeric'
          ? `<label>Min output<input data-i="${i}" data-k="min" value="${escapeHtml(f.min)}"></label><label>Max output<input data-i="${i}" data-k="max" value="${escapeHtml(f.max)}"></label>${f.supports_window ? `<label>Window quarters<input data-i="${i}" data-k="n_quarters" value="${escapeHtml(f.n_quarters)}"></label>` : `<span></span>`}<span></span>`
          : `<label class="wide">Output equals / in list<input data-i="${i}" data-k="text" value="${escapeHtml(f.text)}"></label>`}
        <button class="danger" data-remove="${i}">Remove</button>
      </div>
      <button class="secondary" data-help="${i}" style="margin-top:8px">?</button>
      <div class="definition" id="def-${i}">${escapeHtml(f.definition)}</div>
    </div>`).join('');
  document.querySelectorAll('[data-i]').forEach(el => el.oninput = el.onchange = () => {
    const f = selected[Number(el.dataset.i)], k = el.dataset.k;
    f[k] = el.type === 'checkbox' ? el.checked : el.value;
    refreshQuery();
  });
  document.querySelectorAll('[data-remove]').forEach(btn => btn.onclick = () => { selected.splice(Number(btn.dataset.remove),1); renderMatches(); renderSelected(); refreshQuery(); });
  document.querySelectorAll('[data-help]').forEach(btn => btn.onclick = () => { const d = $('def-'+btn.dataset.help); d.style.display = d.style.display === 'block' ? 'none' : 'block'; });
}

function config(){
  return { geography:$('geography').value, currency:$('currency').value, sort_field:defaultSortField(), sort_direction:'Ascending', fields:selected };
}

async function refreshQuery(){
  const r = await fetch('/query', {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(config())});
  if(!r.ok) $('status').textContent = await r.text();
}

$('search').oninput = renderMatches;
$('add').onclick = () => addField(activeMatch);
$('template').onchange = () => {
  const template = templateByKey($('template').value);
  $('templateDescription').textContent = template?.description || '';
  $('templateName').value = template?.user_template ? template.label : '';
};
$('applyTemplate').onclick = () => applyTemplate($('template').value);
$('saveTemplate').onclick = () => saveTemplate();
$('updateTemplate').onclick = () => saveTemplate({overwrite:true});
$('deleteTemplate').onclick = deleteTemplate;
$('geography').onchange = refreshQuery; $('currency').onchange = refreshQuery;
$('addCustom').onclick = () => {
  const label = $('customLabel').value.trim(), expression = $('customExpression').value.trim();
  if(!expression){ $('status').textContent = 'Custom field needs a BQL expression.'; return; }
  const alias = safeAlias($('customAlias').value || label || expression);
  const supportsWindow = expression.includes('{fpo_start}') || expression.includes('{n_quarters}');
  selected.push({key:'custom_'+Date.now(), label:label || alias, alias, expression, field_type:$('customType').value, definition:$('customDefinition').value || 'Custom Bloomberg/BQL field.', input:true, output:true, min:'', max:'', text:'', n_quarters:12, supports_window:supportsWindow});
  renderSelected(); refreshQuery();
};
async function downloadWorkbook(path, filename){
  $('status').textContent = 'Generating workbook...';
  const r = await fetch(path, {method:'POST', headers:{'Content-Type':'application/json'}, body:JSON.stringify(config())});
  if(!r.ok){ $('status').textContent = await r.text(); return; }
  const blob = await r.blob(); const url = URL.createObjectURL(blob); const a = document.createElement('a');
  a.href = url; a.download = filename; a.click(); URL.revokeObjectURL(url);
  $('status').textContent = 'Workbook downloaded.';
}
$('download').onclick = () => downloadWorkbook('/generate', 'adaptive_screener_workbook.xlsx');
init();
</script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def do_GET(self) -> None:
        path = urllib.parse.urlparse(self.path).path
        if path == "/":
            self.send_bytes(HTML_PAGE.encode("utf-8"), "text/html; charset=utf-8")
        elif path == "/catalog":
            self.send_json(catalog_payload())
        else:
            self.send_error(404, "Not found")

    def do_POST(self) -> None:
        try:
            config = self.read_json()
            if urllib.parse.urlparse(self.path).path == "/query":
                self.send_bytes(build_query(config).encode("utf-8"), "text/plain; charset=utf-8")
            elif urllib.parse.urlparse(self.path).path == "/generate":
                content = build_workbook(config)
                self.send_bytes(
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="adaptive_screener_workbook.xlsx"'},
                )
            elif urllib.parse.urlparse(self.path).path == "/template-save":
                template = upsert_user_template(config)
                self.send_json({"template": template, "user_templates": load_user_templates(), "template_path": str(USER_TEMPLATE_PATH)})
            elif urllib.parse.urlparse(self.path).path == "/template-delete":
                user_templates = delete_user_template(str(config.get("key", "")))
                self.send_json({"user_templates": user_templates, "template_path": str(USER_TEMPLATE_PATH)})
            elif urllib.parse.urlparse(self.path).path == "/clean-pasted":
                raw_text = str(config.get("raw_text", ""))
                content = build_clean_workbook(config, raw_text)
                self.send_bytes(
                    content,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="adaptive_screener_clean_view.xlsx"'},
                )
            else:
                self.send_error(404, "Not found")
        except Exception as exc:
            self.send_response(400)
            self.end_headers()
            self.wfile.write(f"{type(exc).__name__}: {exc}".encode("utf-8"))

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8")) if raw else {}

    def send_json(self, payload: dict) -> None:
        self.send_bytes(json.dumps(payload).encode("utf-8"), "application/json")

    def send_bytes(self, content: bytes, content_type: str, headers: dict[str, str] | None = None) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(content)

    def log_message(self, format: str, *args) -> None:
        print(f"{self.address_string()} - {format % args}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=int(os.environ.get("PORT", "8765")), type=int)
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    print(f"Adaptive Screener Excel Generator running at http://{args.host}:{args.port}")
    server.serve_forever()


if __name__ == "__main__":
    main()
